"""
Task #119 — Mirror the five reconciled rows (A-07, C-11, E-01, E-02, G-01)
into the v4 workbook that the dashboard's status report references.

This is a one-off sync utility paired with `reconcile_feedback_status.py`
(which owns the v3 workbook + markdown report). Do not reuse this for
broader reconciliation work — for that, extend `reconcile_feedback_status.py`
to also write the v4 workbook in one pass.

What each row gets here:
- A-07, E-01, E-02 — status was Open in v4; flip to Done with bilingual
  EN | ES verification notes citing the merged task (#105 / #106).
- C-11 — was already Done in v4 with a bilingual note, but the note
  didn't include the explicit #105 citation; refresh for traceability.
- G-01 — was already Done in v4 but the verification note was English-only;
  refresh to bilingual EN | ES for parity. (Also flips G-01 in V2 Backlog,
  where it was still showing Open.)
"""
import openpyxl
from pathlib import Path

V4 = "attached_assets/reports/KONTi_Dashboard_Feedback_Consolidated_v4.xlsx"

FLIPS = {
    "A-07": (
        "Done",
        "EN: Done in #105 — Photos & Media tab now lives on the project detail "
        "with bulk site-photo upload, category tags, and a per-project gallery "
        "rendered into the project report. File: "
        "artifacts/konti-dashboard/src/components/site-photos-gallery.tsx + "
        "artifacts/konti-dashboard/src/pages/project-detail.tsx. "
        "| ES: Hecho en #105 — la pestaña Fotos & Medios ya existe en el detalle "
        "del proyecto con carga masiva de fotos de obra, etiquetas por categoría "
        "y una galería por proyecto incluida en el reporte. Archivo: "
        "artifacts/konti-dashboard/src/components/site-photos-gallery.tsx + "
        "artifacts/konti-dashboard/src/pages/project-detail.tsx."
    ),
    "E-01": (
        "Done",
        "EN: Done in #106 — Permits page now renders the legal/engineer header "
        "block above the list, mirroring the team's spreadsheet. File: "
        "artifacts/konti-dashboard/src/pages/permits.tsx. "
        "| ES: Hecho en #106 — la página de Permisos ahora muestra el bloque "
        "de encabezado legal/ingeniero arriba del listado, replicando la hoja "
        "de cálculo del equipo. Archivo: "
        "artifacts/konti-dashboard/src/pages/permits.tsx."
    ),
    "E-02": (
        "Done",
        "EN: Done in #106 — Permits are now grouped by type (PCOC, USO, "
        "Consulta de Ubicación, etc.) with separate sections per family, "
        "matching the team's permit Excel. File: "
        "artifacts/konti-dashboard/src/pages/permits.tsx. "
        "| ES: Hecho en #106 — los Permisos ahora están agrupados por tipo "
        "(PCOC, USO, Consulta de Ubicación, etc.) con secciones separadas por "
        "familia, replicando el Excel de permisos del equipo. Archivo: "
        "artifacts/konti-dashboard/src/pages/permits.tsx."
    ),
    # C-11 was already Done in v4 with a bilingual note, but the original
    # note didn't include the explicit #105 task citation that the other
    # reconciled rows have. Refresh it for consistent traceability (Task #119).
    "C-11": (
        "Done",
        "EN: Done in #105 — Site photos in the project report: "
        "site-photos-gallery component is rendered on the project report "
        "and PHOTO_CATEGORY_OPTIONS is wired into the report's photos block "
        "so each photo carries a category label. File: "
        "artifacts/konti-dashboard/src/pages/project-report.tsx + "
        "artifacts/konti-dashboard/src/components/site-photos-gallery.tsx. "
        "| ES: Hecho en #105 — Fotos de obra en el reporte del proyecto: "
        "el componente site-photos-gallery se muestra en el reporte del "
        "proyecto y PHOTO_CATEGORY_OPTIONS está conectado al bloque de "
        "fotos del reporte para que cada foto tenga etiqueta de categoría. "
        "Archivo: artifacts/konti-dashboard/src/pages/project-report.tsx + "
        "artifacts/konti-dashboard/src/components/site-photos-gallery.tsx."
    ),
    # G-01 was already Done in v4 but the verification note was English-only;
    # refresh it to the bilingual EN | ES format for parity with the rest of
    # the reconciled rows (Task #119).
    "G-01": (
        "Done",
        "EN: Already shipped despite V2 scope — ContractorUploadModal "
        "(single + CSV modes) on the Team Directory page. File: "
        "artifacts/konti-dashboard/src/pages/team.tsx (~L69-115). "
        "| ES: Ya implementado a pesar de estar en alcance V2 — "
        "ContractorUploadModal (modos individual + CSV) en la página de "
        "Directorio del Equipo. Archivo: "
        "artifacts/konti-dashboard/src/pages/team.tsx (~L69-115)."
    ),
}

ID_COL = 1
STATUS_COL = 12
VERIF_COL = 15


def flip_sheet(ws):
    flipped = []
    for r in range(2, ws.max_row + 1):
        id_ = ws.cell(r, ID_COL).value
        if not id_:
            continue
        id_ = str(id_).strip()
        if id_ in FLIPS:
            new_status, new_verif = FLIPS[id_]
            old_status = ws.cell(r, STATUS_COL).value
            ws.cell(r, STATUS_COL).value = new_status
            ws.cell(r, VERIF_COL).value = new_verif
            flipped.append((id_, old_status, new_status))
    return flipped


def refresh_summary(ws_summary, ws_main):
    """Update the Summary sheet's status counts based on the current main sheet.

    The Summary sheet contains *three* sections that share the same row labels
    ('Done', 'Open', 'In Progress', 'Needs Decision'):

      1. 'By Status'                       — current totals (always refresh)
      2. 'Audit snapshot (YYYY-MM-DD)'     — today's audit (refresh; the
                                              snapshot date == today's date)
      3. 'Previous snapshot (YYYY-MM-DD)'  — yesterday's audit (NEVER touch;
                                              this is historical data)

    A naive "first match wins" or "last match wins" loop will corrupt the
    historical snapshot. We instead walk the section headers explicitly and
    only update labels that appear after the current header but before the
    next header.
    """
    counts = {}
    for r in range(2, ws_main.max_row + 1):
        s = ws_main.cell(r, STATUS_COL).value
        if s:
            counts[s] = counts.get(s, 0) + 1

    section_starts = []
    for r in range(1, ws_summary.max_row + 1):
        label = ws_summary.cell(r, 1).value
        if isinstance(label, str) and (
            label == "By Status"
            or label.startswith("Audit snapshot")
            or label.startswith("Previous snapshot")
        ):
            section_starts.append((r, label))

    refreshable_starts = [
        (r, lbl) for r, lbl in section_starts
        if lbl == "By Status" or lbl.startswith("Audit snapshot")
    ]

    for idx, (start_row, _label) in enumerate(refreshable_starts):
        next_starts = [
            r for r, _ in section_starts if r > start_row
        ]
        end_row = min(next_starts) if next_starts else ws_summary.max_row + 1

        for r in range(start_row + 1, end_row):
            label = ws_summary.cell(r, 1).value
            if isinstance(label, str) and label in counts:
                ws_summary.cell(r, 2).value = counts[label]

    return counts


PREVIOUS_SNAPSHOT_2026_04_29 = {
    "Done": 16,
    "In Progress": 16,
    "Open": 18,
    "Needs Decision": 7,
}


def restore_previous_snapshot(ws_summary):
    """The first version of refresh_summary() in this helper accidentally
    overwrote the 'Previous snapshot (2026-04-29)' section because it kept
    re-binding labels on every match. Restore it to the original 2026-04-29
    values so the workbook's history is intact again. Idempotent.
    """
    target_header = "Previous snapshot (2026-04-29)"
    header_row = None
    for r in range(1, ws_summary.max_row + 1):
        if ws_summary.cell(r, 1).value == target_header:
            header_row = r
            break
    if header_row is None:
        return False
    for r in range(header_row + 1, ws_summary.max_row + 2):
        label = ws_summary.cell(r, 1).value
        if not isinstance(label, str) or label not in PREVIOUS_SNAPSHOT_2026_04_29:
            if label is None or (isinstance(label, str) and label.strip() == ""):
                continue
            break
        ws_summary.cell(r, 2).value = PREVIOUS_SNAPSHOT_2026_04_29[label]
    return True


def main():
    path = Path(V4)
    wb = openpyxl.load_workbook(path)
    main_ws = wb["KONTi Dashboard Feedback"]
    flipped_main = flip_sheet(main_ws)
    print(f"Main sheet flips: {flipped_main}")

    if "V2 Backlog" in wb.sheetnames:
        flipped_v2 = flip_sheet(wb["V2 Backlog"])
        print(f"V2 Backlog flips: {flipped_v2}")

    if "Summary" in wb.sheetnames:
        counts = refresh_summary(wb["Summary"], main_ws)
        print(f"Summary counts: {counts}")
        if restore_previous_snapshot(wb["Summary"]):
            print(f"Restored 'Previous snapshot (2026-04-29)' to "
                  f"{PREVIOUS_SNAPSHOT_2026_04_29}")

    wb.save(path)
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
